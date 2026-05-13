"""Excel 작업 명세서를 JSON 으로 변환.

레이아웃: A열 = 항목 키, B열 = 값. 시트 첫 장 사용.

지원 키 (대소문자/공백 무관, 미설정 키는 출력에서 생략):
    source_font, output_path
    family_en, family_ko, subfamily_en, subfamily_ko, full_name_en
    version, copyright, manufacturer, designer, license
    postscript_name
    ascender, descender, line_gap

사용:  python parse_spec.py <spec.xlsx>
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


# 키 정규화 → 표준 이름
_KEY_ALIASES = {
    "source": "source_font", "input": "source_font", "font": "source_font",
    "source_font": "source_font", "src": "source_font",
    "output": "output_path", "output_path": "output_path", "out": "output_path",
    "family": "family_en", "family_en": "family_en", "family_english": "family_en",
    "family_ko": "family_ko", "family_korean": "family_ko", "korean_family": "family_ko",
    "subfamily": "subfamily_en", "subfamily_en": "subfamily_en", "style": "subfamily_en",
    "subfamily_ko": "subfamily_ko",
    "full_name": "full_name_en", "fullname": "full_name_en", "full_name_en": "full_name_en",
    "version": "version",
    "copyright": "copyright",
    "manufacturer": "manufacturer", "vendor": "manufacturer",
    "designer": "designer",
    "license": "license", "license_description": "license",
    "postscript_name": "postscript_name", "postscriptname": "postscript_name", "ps_name": "postscript_name",
    "ascender": "ascender", "ascent": "ascender",
    "descender": "descender", "descent": "descender",
    "line_gap": "line_gap", "linegap": "line_gap", "leading": "line_gap",
}

_INT_KEYS = {"ascender", "descender", "line_gap"}


def _norm(key: str) -> str | None:
    k = str(key).strip().lower().replace(" ", "_").replace("-", "_")
    return _KEY_ALIASES.get(k)


def parse(xlsx_path: str) -> dict:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    spec: dict[str, object] = {}
    unknown: list[str] = []

    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        if row[0] is None:
            continue
        raw_key = str(row[0])
        value = row[1]
        if value is None or str(value).strip() == "":
            continue

        std = _norm(raw_key)
        if std is None:
            unknown.append(raw_key)
            continue

        if std in _INT_KEYS:
            try:
                value = int(value)
            except (TypeError, ValueError):
                raise ValueError(f"{raw_key} must be integer, got {value!r}")
        else:
            value = str(value).strip()

        spec[std] = value

    return {"spec": spec, "unknown_keys": unknown, "sheet": ws.title}


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("usage: python parse_spec.py <spec.xlsx>", file=sys.stderr)
        sys.exit(2)
    path = Path(sys.argv[1])
    if not path.exists():
        print(f"file not found: {path}", file=sys.stderr)
        sys.exit(2)
    print(json.dumps(parse(str(path)), indent=2, ensure_ascii=False))
