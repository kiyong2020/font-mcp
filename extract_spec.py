"""기존 폰트 패밀리 → 6-시트 작업서(XLSX) 추출기.

`scripts/make_build_template.py` 가 만드는 빈 템플릿과 동일한 시트 구조에
실제 폰트의 name 테이블·메트릭 값을 채워 내보낸다.

쓰임:
    spec = extract_spec.extract(
        font_paths=["/abs/Rg.otf", "/abs/Bd.otf", ...],
        output_path="/abs/work_order.xlsx",
    )

설계:
- MCP / font_ops 의존성 없음. 입력은 폰트 경로 리스트, 출력은 XLSX.
- 메타·메트릭은 Regular(usWeightClass=400) 우선, 없으면 첫 번째 폰트 기준.
- weights 시트는 usWeightClass 오름차순으로 한 행씩.
- names 시트는 build_font_family 가 자동 생성하지 않는 nameID 만 (0/7/8/9/11~14 등).
"""
from __future__ import annotations

from pathlib import Path
from typing import Optional

from fontTools.ttLib import TTFont
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# fsSelection / macStyle 비트 (font_ops 의 상수와 동일)
_USE_TYPO_METRICS = 1 << 7
_HEAD_BOLD = 1 << 0
_HEAD_ITALIC = 1 << 1

# build_font_family 가 weight/meta 로부터 자동 생성하는 nameID — names 시트에서는 제외
_AUTO_GENERATED_NAME_IDS = {1, 2, 3, 4, 5, 6, 16, 17, 18}

# Windows langID ↔ 시트 약어
_LANG_REV = {
    0x0409: "en",
    0x0412: "ko",
    0x0411: "ja",
    0x0804: "zh-cn",
    0x0404: "zh-tw",
}
_PLATFORM_REV = {
    (3, 1): "win",
    (1, 0): "mac",
}

# 시트 스타일 — make_build_template.py 와 동일한 룩앤필 유지
_HEADER_FILL = PatternFill("solid", fgColor="2F5B7C")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_SECTION_FILL = PatternFill("solid", fgColor="E8EEF4")
_SECTION_FONT = Font(bold=True)


def _set_header(ws, row: int, values: list[str]) -> None:
    for col, val in enumerate(values, start=1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center")


def _autosize(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── name 테이블 헬퍼 ──────────────────────────────────────────────

def _name(font: TTFont, nid: int, lang_id: int = 0x0409) -> str:
    """nameID 조회. (platformID=3, encID=1, langID) 우선, 없으면 같은 nameID 의 첫 매치."""
    rec = font["name"].getName(nid, 3, 1, lang_id)
    if rec is not None:
        return rec.toUnicode()
    # platform=3 의 다른 langID
    for r in font["name"].names:
        if r.nameID == nid and r.platformID == 3:
            return r.toUnicode()
    # 그 외 platform
    for r in font["name"].names:
        if r.nameID == nid:
            return r.toUnicode()
    return ""


def _pick_base(fonts: list[TTFont]) -> TTFont:
    """meta/metrics 추출 기준 폰트 — Regular(400) 우선."""
    for f in fonts:
        if f["OS/2"].usWeightClass == 400:
            return f
    return fonts[0]


def _detect_scope(fonts: list[TTFont]) -> str:
    """ko-KR 레코드가 하나라도 있으면 'korean', 아니면 'latin'."""
    for f in fonts:
        for r in f["name"].names:
            if r.langID == 0x0412:
                return "korean"
    return "latin"


def _strikeout(font: TTFont) -> tuple[Optional[int], Optional[int]]:
    os2 = font["OS/2"]
    pos = getattr(os2, "yStrikeoutPosition", None)
    sz = getattr(os2, "yStrikeoutSize", None)
    return pos, sz


def _underline(font: TTFont) -> tuple[Optional[int], Optional[int]]:
    post = font.get("post")
    if post is None:
        return None, None
    return (
        getattr(post, "underlinePosition", None),
        getattr(post, "underlineThickness", None),
    )


# ── 시트 빌더 ────────────────────────────────────────────────────

def _meta_sheet(wb: Workbook, base: TTFont, scope: str) -> None:
    ws = wb.create_sheet("meta")
    _set_header(ws, 1, ["key", "value", "notes"])

    head = base["head"]
    rows = [
        ("project_name", _name(base, 1) or _name(base, 16),
         "프로젝트 이름 — nameID 1 (family) 기준"),
        ("version", f"{head.fontRevision:.3f}", "head.fontRevision"),
        ("author", _name(base, 9), "작업자 — nameID 9 (en)"),
        ("client", "", "고객사 (자동 추출 불가, 직접 입력)"),
        ("vendor", _name(base, 8), "제작사 — nameID 8"),
        ("designer_ko", _name(base, 9, 0x0412), "디자이너 한글명 — nameID 9 ko"),
        ("designer_en", _name(base, 9, 0x0409), "디자이너 영문명 — nameID 9 en"),
        ("vendor_url", _name(base, 11), "nameID 11"),
        ("designer_url", _name(base, 12), "nameID 12"),
        ("license", _name(base, 13), "nameID 13"),
        ("license_url", _name(base, 14), "nameID 14"),
        ("copyright", _name(base, 0), "nameID 0"),
        ("trademark", _name(base, 7), "nameID 7"),
        ("name_record_scope", scope, "korean | latin — ko-KR 레코드 유무로 추정"),
    ]
    for i, row in enumerate(rows, start=2):
        for col, val in enumerate(row, start=1):
            ws.cell(row=i, column=col, value=val)
    _autosize(ws, [22, 60, 50])


def _metrics_sheet(wb: Workbook, base: TTFont) -> None:
    ws = wb.create_sheet("metrics")
    _set_header(ws, 1, ["table", "field", "value", "notes"])

    head = base["head"]
    hhea = base["hhea"]
    os2 = base["OS/2"]
    strk_pos, strk_sz = _strikeout(base)
    ul_pos, ul_thk = _underline(base)
    use_typo = bool(os2.fsSelection & _USE_TYPO_METRICS)

    rows = [
        ("head", "unitsPerEm", head.unitsPerEm, "UPM"),
        ("hhea", "ascent", hhea.ascent, ""),
        ("hhea", "descent", hhea.descent, "음수 일반적"),
        ("hhea", "lineGap", hhea.lineGap, ""),
        ("OS/2", "sTypoAscender", os2.sTypoAscender, ""),
        ("OS/2", "sTypoDescender", os2.sTypoDescender, "음수"),
        ("OS/2", "sTypoLineGap", os2.sTypoLineGap, ""),
        ("OS/2", "usWinAscent", os2.usWinAscent, "양수"),
        ("OS/2", "usWinDescent", os2.usWinDescent, "양수"),
        ("OS/2", "yStrikeoutPosition", strk_pos if strk_pos is not None else "", ""),
        ("OS/2", "yStrikeoutSize", strk_sz if strk_sz is not None else "", ""),
        ("OS/2", "fsType", os2.fsType,
         "0=Installable, 2=Restricted, 4=Preview&Print, 8=Editable"),
        ("post", "underlinePosition", ul_pos if ul_pos is not None else "", "음수"),
        ("post", "underlineThickness", ul_thk if ul_thk is not None else "", ""),
        ("OS/2", "use_typo_metrics", "TRUE" if use_typo else "FALSE",
         "fsSelection bit 7 — 행간 일관성"),
    ]
    for i, row in enumerate(rows, start=2):
        for col, val in enumerate(row, start=1):
            ws.cell(row=i, column=col, value=val)
    _autosize(ws, [10, 22, 14, 60])


def _weight_row(font: TTFont, source_path: str) -> tuple:
    """폰트 1개 → weights 시트 한 행."""
    os2 = font["OS/2"]
    head = font["head"]
    weight_class = os2.usWeightClass
    style_name = _name(font, 2) or _name(font, 17)
    is_bold = bool(head.macStyle & _HEAD_BOLD)
    is_italic = bool(head.macStyle & _HEAD_ITALIC)
    korean_family = ""
    # platform=3, lang=ko 의 nameID 16(typo family) 우선, 없으면 1
    rec = font["name"].getName(16, 3, 1, 0x0412) or font["name"].getName(1, 3, 1, 0x0412)
    if rec is not None:
        korean_family = rec.toUnicode()
    latin_family = _name(font, 16, 0x0409) or _name(font, 1, 0x0409)
    psname = _name(font, 6)
    full_name = _name(font, 4) or _name(font, 18)
    # fullname_suffix: full name 에서 family prefix 를 떼낸 나머지
    fullname_suffix = style_name
    if latin_family and full_name.startswith(latin_family):
        suffix = full_name[len(latin_family):].strip()
        if suffix:
            fullname_suffix = suffix
    base_font_key = Path(source_path).stem
    return (
        weight_class, style_name, "TRUE" if is_bold else "FALSE",
        "TRUE" if is_italic else "FALSE",
        korean_family, latin_family, psname, fullname_suffix, base_font_key, "",
    )


def _weights_sheet(wb: Workbook, fonts: list[TTFont], paths: list[str]) -> None:
    ws = wb.create_sheet("weights")
    _set_header(
        ws, 1,
        [
            "weight_class", "style_name", "is_bold", "is_italic",
            "korean_family", "latin_family", "psname", "fullname_suffix",
            "base_font_key", "notes",
        ],
    )
    pairs = sorted(zip(fonts, paths), key=lambda x: x[0]["OS/2"].usWeightClass)
    for i, (f, p) in enumerate(pairs, start=2):
        row = _weight_row(f, p)
        for col, val in enumerate(row, start=1):
            ws.cell(row=i, column=col, value=val)
    _autosize(ws, [12, 12, 8, 9, 28, 30, 18, 16, 18, 30])


def _outputs_sheet(wb: Workbook, paths: list[str], fonts: list[TTFont]) -> None:
    ws = wb.create_sheet("outputs")
    _set_header(ws, 1, ["format", "enabled", "psname_suffix", "notes"])

    exts = {Path(p).suffix.lower() for p in paths}
    has_otf = ".otf" in exts
    has_ttf = ".ttf" in exts
    has_vf = any("fvar" in f for f in fonts)

    rows = [
        ("OTF", "TRUE" if has_otf else "FALSE", "OTF",
         "OpenType CFF — 베이스에 .otf 가 있어야 함"),
        ("TTF", "TRUE" if has_ttf else "FALSE", "",
         "TrueType — 베이스가 .ttf 라야 함"),
        ("WOFF", "FALSE", "", "WOFF1 래핑"),
        ("WOFF2", "TRUE", "", "WOFF2 래핑"),
        ("WOFF_subset", "FALSE", "", "subset 기준 서브셋팅 후 WOFF1"),
        ("WOFF2_subset", "TRUE", "", "subset 기준 서브셋팅 후 WOFF2"),
        ("VF", "TRUE" if has_vf else "FALSE", "VF",
         "variable_font 인자 필수. 입력에 fvar 가 있으면 자동 ON"),
    ]
    for i, row in enumerate(rows, start=2):
        for col, val in enumerate(row, start=1):
            ws.cell(row=i, column=col, value=val)
    _autosize(ws, [16, 10, 16, 60])


def _names_sheet(wb: Workbook, fonts: list[TTFont]) -> None:
    ws = wb.create_sheet("names")
    _set_header(ws, 1, ["nameID", "platform", "lang", "value", "notes"])

    # 모든 폰트에서 auto-generated 가 아닌 nameID 를 모은다.
    # (nameID, platformID, platEncID, langID, value) 5-튜플로 중복 제거.
    seen: set[tuple] = set()
    rows: list[tuple] = []
    for f in fonts:
        for r in f["name"].names:
            if r.nameID in _AUTO_GENERATED_NAME_IDS:
                continue
            plat_key = (r.platformID, r.platEncID)
            if plat_key not in _PLATFORM_REV:
                continue
            if r.langID not in _LANG_REV:
                continue
            try:
                value = r.toUnicode()
            except Exception:
                continue
            if not value.strip():
                continue
            key = (r.nameID, r.platformID, r.platEncID, r.langID, value)
            if key in seen:
                continue
            seen.add(key)
            rows.append((
                r.nameID,
                _PLATFORM_REV[plat_key],
                _LANG_REV[r.langID],
                value,
                "",
            ))
    rows.sort(key=lambda x: (x[0], x[1], x[2]))

    for i, row in enumerate(rows, start=2):
        for col, val in enumerate(row, start=1):
            ws.cell(row=i, column=col, value=val)
    ws.cell(
        row=len(rows) + 3, column=1,
        value="* nameID 1,2,3,4,5,6,16,17 은 weights+meta 기반으로 빌드 시 자동 생성됨 (여기 비워둬도 됨)",
    )
    _autosize(ws, [10, 10, 8, 70, 40])


def _subset_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("subset")
    _set_header(ws, 1, ["key", "value", "notes"])
    rows = [
        ("mode", "preset", "text | unicodes | preset"),
        ("text", "", "유지할 글자들 (mode=text)"),
        ("unicodes", "",
         "콤마/공백 구분 (예: 0x20-0x7E, 0xAC00-0xD7A3)"),
        ("preset", "common_kr",
         "common_kr: ASCII + 한글 음절 + 자주 쓰는 약물"),
        ("layout_features", "*",
         "* = 모든 GSUB/GPOS. 아니면 'kern,liga' 식"),
    ]
    for i, row in enumerate(rows, start=2):
        for col, val in enumerate(row, start=1):
            ws.cell(row=i, column=col, value=val)
    _autosize(ws, [18, 28, 60])


def _readme_sheet(wb: Workbook, font_paths: list[str]) -> None:
    ws = wb.create_sheet("README", 0)
    ws["A1"] = "[제작의뢰서] 작업서 — 기존 패밀리에서 자동 추출"
    ws["A1"].font = Font(bold=True, size=14)
    lines = [
        "",
        "이 워크북은 font-mcp 의 build_font_family 가 읽는 입력 포맷입니다.",
        "기존 폰트의 name 테이블·메트릭을 그대로 채워 넣은 상태이므로,",
        "필요한 값(프로젝트명/저작권/디자이너 등)만 검토·수정해서 사용하면 됩니다.",
        "",
        "■ 입력 폰트",
    ]
    for p in font_paths:
        lines.append(f"  - {p}")
    lines += [
        "",
        "■ 시트 구성",
        "  meta      — 프로젝트/저작권/벤더 (nameID 0/7/8/9/11/12/13/14)",
        "  metrics   — head·hhea·OS/2·post 메트릭",
        "  weights   — 각 웨이트의 nameID 1/2/4/6 + macStyle 비트",
        "  outputs   — 입력 확장자 기반 권장값 (.otf→OTF, .ttf→TTF, fvar→VF)",
        "  names     — auto-generated 외의 nameID 레코드",
        "  subset    — WOFF*_subset 빌드 규칙 (기본값)",
        "",
        "■ 다음 단계",
        "  1. Google Drive 에 업로드 → Sheets 로 값 검토",
        "  2. 파일 → 다운로드 → Microsoft Excel(.xlsx)",
        "  3. Claude 에게: 'sheet_path=…, base_otf={…} 로 빌드해줘'",
    ]
    for i, line in enumerate(lines, start=2):
        c = ws.cell(row=i, column=1, value=line)
        if line.startswith("■"):
            c.font = _SECTION_FONT
            c.fill = _SECTION_FILL
    ws.column_dimensions["A"].width = 110


# ── 공개 API ─────────────────────────────────────────────────────

def extract(font_paths: list[str], output_path: str) -> dict:
    """폰트 패밀리 → 채워진 6-시트 XLSX 작업서.

    Args:
        font_paths: 패밀리에 속한 폰트들의 절대 경로. 1개 이상.
        output_path: 결과 XLSX 경로.

    Returns:
        {output_path, font_count, weight_classes, name_record_scope, base_font}
    """
    if not font_paths:
        raise ValueError("font_paths 가 비어있음")
    fonts: list[TTFont] = []
    for p in font_paths:
        if not Path(p).exists():
            raise FileNotFoundError(p)
        fonts.append(TTFont(p))

    base = _pick_base(fonts)
    scope = _detect_scope(fonts)
    base_idx = fonts.index(base)

    wb = Workbook()
    wb.remove(wb.active)
    _readme_sheet(wb, font_paths)
    _meta_sheet(wb, base, scope)
    _metrics_sheet(wb, base)
    _weights_sheet(wb, fonts, font_paths)
    _outputs_sheet(wb, font_paths, fonts)
    _names_sheet(wb, fonts)
    _subset_sheet(wb)

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    return {
        "output_path": str(out),
        "font_count": len(fonts),
        "weight_classes": sorted({f["OS/2"].usWeightClass for f in fonts}),
        "name_record_scope": scope,
        "base_font": str(font_paths[base_idx]),
    }
