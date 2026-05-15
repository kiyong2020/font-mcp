"""Google Sheets 호환 XLSX 작업서 템플릿 생성기.

[제작의뢰서] Confluence 페이지(예: 4440719397 현대캐피탈 산스)의
스키마를 그대로 옮긴 6-시트 XLSX 를 만든다.

Usage:
    python scripts/make_build_template.py [output.xlsx]

기본 출력: ./build_template.xlsx (현대캐피탈 산스 샘플값 포함)

업로드 흐름:
    1) 이 스크립트로 만든 XLSX 를 Google Drive 에 업로드 → Google Sheets 로 열기
    2) 디자이너가 값 수정
    3) 시트 → 파일 → 다운로드 → Microsoft Excel(.xlsx)
    4) MCP build_font_family(sheet_path=…) 로 입력
"""
from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="2F5B7C")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="E8EEF4")
SECTION_FONT = Font(bold=True)


def _set_header(ws, row: int, values: list[str]) -> None:
    for col, val in enumerate(values, start=1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center")


def _autosize(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _meta_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("meta")
    _set_header(ws, 1, ["key", "value", "notes"])
    rows = [
        ("project_name", "현대캐피탈 산스", "프로젝트 이름 (자유 입력)"),
        ("version", "1.000", "fontRevision (head 테이블)"),
        ("author", "권경석", "작업자 (메모용)"),
        ("client", "HYUNDAI CAPITAL SERVICES, Inc.", "고객사"),
        ("vendor", "Sandoll Inc.", "제작사 — nameID=8"),
        ("designer_ko", "권경석", "디자이너 한글명 — nameID=9 ko"),
        ("designer_en", "Kweon Kyoung-Seok", "디자이너 영문명 — nameID=9 en"),
        ("vendor_url", "https://www.sandoll.co.kr/", "nameID=11"),
        ("designer_url", "https://www.sandoll.co.kr/", "nameID=12"),
        ("license", "HYUNDAI CAPITAL SERVICES, Inc.", "nameID=13"),
        ("license_url", "https://www.hyundaicapital.com", "nameID=14"),
        ("copyright",
         "COPYRIGHT (c) HYUNDAI CAPITAL SERVICES, Inc. ALL RIGHTS RESERVED.",
         "nameID=0"),
        ("trademark", "HCS Sans is a trademark of HCS.", "nameID=7"),
        ("name_record_scope", "korean", "korean | latin — 인스톨러/패키지 구분"),
    ]
    for i, row in enumerate(rows, start=2):
        for col, val in enumerate(row, start=1):
            ws.cell(row=i, column=col, value=val)
    _autosize(ws, [22, 60, 50])


def _metrics_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("metrics")
    _set_header(ws, 1, ["table", "field", "value", "notes"])
    rows = [
        ("head", "unitsPerEm", 1000, "UPM"),
        ("hhea", "ascent", 800, "= hheaAscender"),
        ("hhea", "descent", -300, "= hheaDescender (음수)"),
        ("hhea", "lineGap", 0, ""),
        ("OS/2", "sTypoAscender", 800, ""),
        ("OS/2", "sTypoDescender", -300, "음수"),
        ("OS/2", "sTypoLineGap", 0, ""),
        ("OS/2", "usWinAscent", 800, "양수 (절댓값 권장)"),
        ("OS/2", "usWinDescent", 300, "양수"),
        ("OS/2", "yStrikeoutPosition", 300, ""),
        ("OS/2", "yStrikeoutSize", 50, ""),
        ("OS/2", "fsType", 0,
         "0=Installable, 2=Restricted, 4=Preview&Print, 8=Editable"),
        ("post", "underlinePosition", -100, "음수"),
        ("post", "underlineThickness", 50, ""),
        ("OS/2", "use_typo_metrics", "TRUE",
         "TRUE 면 fsSelection bit7 ON — 행간 일관성"),
    ]
    for i, row in enumerate(rows, start=2):
        for col, val in enumerate(row, start=1):
            ws.cell(row=i, column=col, value=val)
    _autosize(ws, [10, 22, 14, 60])


def _weights_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("weights")
    _set_header(
        ws, 1,
        [
            "weight_class", "style_name", "is_bold", "is_italic",
            "korean_family", "latin_family", "psname", "fullname_suffix",
            "base_font_key", "notes",
        ],
    )
    rows = [
        (800, "ExtraBold", "FALSE", "FALSE",
         "현대캐피탈 산스 Head", "HCS Sans Head", "HCSSansHdEB",
         "ExtraBold", "HdEB", ""),
        (700, "Bold", "TRUE", "FALSE",
         "현대캐피탈 산스 Head Bold", "HCS Sans Head Bold", "HCSSansHdBd",
         "Bold", "HdBd", "macStyle.BOLD on"),
        (600, "SemiBold", "FALSE", "FALSE",
         "현대캐피탈 산스 Head SemiBold", "HCS Sans Head SemiBold",
         "HCSSansHdSB", "SemiBold", "HdSB", ""),
        (500, "Medium", "FALSE", "FALSE",
         "현대캐피탈 산스 Text", "HCS Sans Text", "HCSSansTxMd",
         "Medium", "TxMd", ""),
        (400, "Regular", "FALSE", "FALSE",
         "현대캐피탈 산스 Text", "HCS Sans Text", "HCSSansTxRg",
         "Regular", "TxRg", "fsSelection.REGULAR on (default)"),
        (300, "Light", "FALSE", "FALSE",
         "현대캐피탈 산스 Text Light", "HCS Sans Text Light",
         "HCSSansTxLt", "Light", "TxLt", ""),
    ]
    for i, row in enumerate(rows, start=2):
        for col, val in enumerate(row, start=1):
            ws.cell(row=i, column=col, value=val)
    _autosize(ws, [12, 12, 8, 9, 28, 30, 18, 16, 16, 30])


def _outputs_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("outputs")
    _set_header(ws, 1, ["format", "enabled", "psname_suffix", "notes"])
    rows = [
        ("OTF", "TRUE", "OTF", "OpenType CFF, PSname 뒤에 'OTF' 붙음"),
        ("TTF", "TRUE", "", "TrueType (베이스가 .ttf 라야 함)"),
        ("WOFF", "FALSE", "", "WOFF1 래핑"),
        ("WOFF2", "TRUE", "", "WOFF2 래핑"),
        ("WOFF_subset", "FALSE", "", "subset 시트 기준 서브셋팅 후 WOFF1"),
        ("WOFF2_subset", "TRUE", "", "subset 시트 기준 서브셋팅 후 WOFF2"),
        ("VF", "TRUE", "VF", "variable_font 인자 필수. 인스턴스 자동 생성 X"),
    ]
    for i, row in enumerate(rows, start=2):
        for col, val in enumerate(row, start=1):
            ws.cell(row=i, column=col, value=val)
    _autosize(ws, [16, 10, 16, 60])


def _names_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("names")
    _set_header(
        ws, 1,
        ["nameID", "platform", "lang", "value", "notes"],
    )
    rows = [
        (0, "win", "en",
         "COPYRIGHT (c) HYUNDAI CAPITAL SERVICES, Inc. ALL RIGHTS RESERVED.",
         "meta.copyright 와 동일하면 비워둬도 됨"),
        (7, "win", "en", "HCS Sans is a trademark of HCS.", ""),
        (8, "win", "en", "Sandoll Inc.", "제작사"),
        (9, "win", "en", "Kweon Kyoung-Seok", "디자이너 영문"),
        (9, "win", "ko", "권경석", "디자이너 한글"),
        (11, "win", "en", "https://www.sandoll.co.kr/", "벤더 URL"),
        (12, "win", "en", "https://www.sandoll.co.kr/", "디자이너 URL"),
        (13, "win", "en", "HYUNDAI CAPITAL SERVICES, Inc.", "라이선스"),
        (14, "win", "en", "https://www.hyundaicapital.com", "라이선스 URL"),
    ]
    for i, row in enumerate(rows, start=2):
        for col, val in enumerate(row, start=1):
            ws.cell(row=i, column=col, value=val)
    ws.cell(
        row=len(rows) + 3, column=1,
        value="* nameID 1,2,3,4,6,16,17 은 weights 시트 + meta 기반으로 자동 생성됨",
    )
    _autosize(ws, [10, 10, 8, 70, 40])


def _subset_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("subset")
    _set_header(ws, 1, ["key", "value", "notes"])
    rows = [
        ("mode", "preset", "text | unicodes | preset 중 하나"),
        ("text", "",
         "유지할 글자들을 한 셀에 모두 (mode=text 일 때만)"),
        ("unicodes", "",
         "콤마/공백 구분 유니코드 (예: 0x20-0x7E, 0xAC00-0xD7A3)"),
        ("preset", "common_kr",
         "common_kr: ASCII + 한글 2350자 + 자주 쓰는 약물"),
        ("layout_features", "*",
         "* = 모든 GSUB/GPOS 유지. 아니면 'kern,liga' 식 콤마 구분"),
    ]
    for i, row in enumerate(rows, start=2):
        for col, val in enumerate(row, start=1):
            ws.cell(row=i, column=col, value=val)
    _autosize(ws, [18, 28, 60])


def _readme_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("README", 0)
    ws["A1"] = "[제작의뢰서 → 폰트 빌드] 작업서"
    ws["A1"].font = Font(bold=True, size=14)
    lines = [
        "",
        "이 워크북은 font-mcp 의 build_font_family 도구가 읽는 입력 포맷입니다.",
        "Confluence [제작의뢰서] 페이지의 스키마를 그대로 따릅니다.",
        "",
        "■ 시트 구성",
        "  meta      — 프로젝트/저작권/벤더 정보 (nameID 0/7/8/9/11/12/13/14)",
        "  metrics   — head·hhea·OS/2·post 테이블 메트릭 값",
        "  weights   — 패밀리 안의 각 웨이트 (한글명/영문명/PSname/WeightClass)",
        "  outputs   — 어떤 포맷을 생성할지 (OTF/TTF/VF/WOFF/WOFF2/subset)",
        "  names     — 추가 nameID 레코드 (커스텀)",
        "  subset    — WOFF*_subset 빌드 시 서브셋팅 규칙",
        "",
        "■ 사용 흐름",
        "  1. 이 XLSX 를 Google Drive 에 올려 시트로 열고 값 수정",
        "  2. 파일 → 다운로드 → Microsoft Excel(.xlsx)",
        "  3. Claude 에게: 'sheet_path=…, base_fonts={400:Rg.otf,…} 로 빌드해줘'",
        "",
        "■ base_fonts 인자",
        "  build_font_family(base_fonts={'400':'/abs/HCSSansTxRg.otf','700':'…'})",
        "  weights 시트의 weight_class 와 키가 일치해야 합니다.",
        "  VF 가 enabled 면 variable_font='/abs/HCSSans-VF.ttf' 추가 전달.",
        "",
        "■ 안전 장치",
        "  - 원본 폰트는 절대 덮어쓰지 않습니다.",
        "  - 출력마다 fontbakery 검증 + cases.json 학습 메모리 자동 기록.",
    ]
    for i, line in enumerate(lines, start=2):
        c = ws.cell(row=i, column=1, value=line)
        if line.startswith("■"):
            c.font = SECTION_FONT
            c.fill = SECTION_FILL
    ws.column_dimensions["A"].width = 110


def build(out_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    _readme_sheet(wb)
    _meta_sheet(wb)
    _metrics_sheet(wb)
    _weights_sheet(wb)
    _outputs_sheet(wb)
    _names_sheet(wb)
    _subset_sheet(wb)
    wb.save(out_path)
    print(f"wrote {out_path}")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "output", nargs="?",
        default=str(
            Path(__file__).resolve().parent.parent / "build_template.xlsx"
        ),
    )
    args = ap.parse_args()
    build(Path(args.output))


if __name__ == "__main__":
    main()
