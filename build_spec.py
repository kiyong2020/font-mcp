"""작업서 XLSX 파서.

scripts/make_build_template.py 가 만드는 6-시트 워크북을 BuildSpec 으로 읽는다.
시트 구조는 동일 스크립트의 docstring 참고.

설계:
- 모든 셀 값은 normalize 단계에서 타입 강제. (예: "TRUE"/"FALSE" → bool)
- 누락/오타는 ValueError 로 즉시 실패. MCP 가 그대로 에이전트에 노출.
- BuildSpec 은 .to_dict() 로 JSON 직렬화 가능.
"""
from __future__ import annotations

from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook


# ── 프리셋: 자주 쓰는 한글 + ASCII 서브셋 ─────────────────────────
# KS X 1001 의 한글 음절 2350 자 코드포인트 — fontTools 의 unicodes 인자에 그대로.
# 정확한 2350 자 리스트는 fontTools 의 KS_X_1001 에서 가져오는 것이 이상적이나,
# 여기서는 가장 자주 쓰는 범위 근사치 (한글 음절 영역 AC00-D7A3 전체) 를 쓴다.
# 디자이너가 정확한 2350 만 원하면 subset 시트에 unicodes 모드로 직접 명시.
_PRESET_COMMON_KR_RANGES: list[tuple[int, int]] = [
    (0x0020, 0x007E),  # ASCII printable
    (0x00A0, 0x00FF),  # Latin-1 supplement
    (0x2000, 0x206F),  # General Punctuation
    (0x2100, 0x214F),  # Letterlike Symbols
    (0x2190, 0x21FF),  # Arrows
    (0x2200, 0x22FF),  # Math operators
    (0x2500, 0x257F),  # Box drawing
    (0x25A0, 0x25FF),  # Geometric shapes
    (0x3000, 0x303F),  # CJK symbols & punctuation
    (0x3130, 0x318F),  # Hangul compat jamo
    (0xAC00, 0xD7A3),  # Hangul syllables (전체 11172)
    (0xFF00, 0xFFEF),  # Halfwidth & Fullwidth forms
]


def _expand_preset(name: str) -> list[int]:
    if name == "common_kr":
        out: list[int] = []
        for lo, hi in _PRESET_COMMON_KR_RANGES:
            out.extend(range(lo, hi + 1))
        return out
    raise ValueError(f"unknown subset preset: {name!r}")


def _parse_unicode_ranges(spec: str) -> list[int]:
    """'0x20-0x7E, 0xAC00-0xD7A3' 같은 표기를 코드포인트 리스트로 전개."""
    out: list[int] = []
    for chunk in spec.replace(";", ",").split(","):
        chunk = chunk.strip()
        if not chunk:
            continue
        if "-" in chunk:
            lo_s, hi_s = chunk.split("-", 1)
            lo = int(lo_s, 16) if lo_s.lower().startswith("0x") else int(lo_s)
            hi = int(hi_s, 16) if hi_s.lower().startswith("0x") else int(hi_s)
            out.extend(range(lo, hi + 1))
        else:
            v = int(chunk, 16) if chunk.lower().startswith("0x") else int(chunk)
            out.append(v)
    return out


def _to_bool(v) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    s = str(v).strip().lower()
    if s in ("true", "1", "yes", "y", "on"):
        return True
    if s in ("false", "0", "no", "n", "off", ""):
        return False
    raise ValueError(f"cannot interpret {v!r} as boolean")


def _to_int(v, *, field: str) -> int:
    if isinstance(v, bool):
        raise ValueError(f"{field}: expected int, got bool")
    if isinstance(v, int):
        return v
    if isinstance(v, float) and v.is_integer():
        return int(v)
    try:
        return int(str(v).strip())
    except Exception as e:
        raise ValueError(f"{field}: cannot parse {v!r} as int") from e


# 언어 약어 → Windows langID
_LANG_MAP = {
    "en": 0x0409,
    "en-us": 0x0409,
    "ko": 0x0412,
    "ko-kr": 0x0412,
    "ja": 0x0411,
    "zh-cn": 0x0804,
    "zh-tw": 0x0404,
}

_PLATFORM_MAP = {
    "win": (3, 1),
    "windows": (3, 1),
    "mac": (1, 0),
}


# ── 데이터 클래스 ──────────────────────────────────────────────

@dataclass
class Meta:
    project_name: str = ""
    version: str = "1.000"
    author: str = ""
    client: str = ""
    vendor: str = ""
    designer_ko: str = ""
    designer_en: str = ""
    vendor_url: str = ""
    designer_url: str = ""
    license: str = ""
    license_url: str = ""
    copyright: str = ""
    trademark: str = ""
    name_record_scope: str = "korean"  # 'korean' | 'latin'


@dataclass
class Metrics:
    units_per_em: int = 1000
    hhea_ascent: int = 800
    hhea_descent: int = -300
    hhea_line_gap: int = 0
    typo_ascender: int = 800
    typo_descender: int = -300
    typo_line_gap: int = 0
    win_ascent: int = 800
    win_descent: int = 300
    strikeout_position: Optional[int] = None
    strikeout_size: Optional[int] = None
    fs_type: int = 0
    underline_position: Optional[int] = None
    underline_thickness: Optional[int] = None
    use_typo_metrics: bool = True


@dataclass
class WeightSpec:
    weight_class: int
    style_name: str
    is_bold: bool
    is_italic: bool
    korean_family: str
    latin_family: str
    psname: str
    fullname_suffix: str
    base_font_key: str  # base_fonts dict 의 키 (보통 str(weight_class))
    notes: str = ""


@dataclass
class OutputSpec:
    fmt: str          # OTF | TTF | WOFF | WOFF2 | WOFF_subset | WOFF2_subset | VF
    enabled: bool
    psname_suffix: str = ""


@dataclass
class NameRecord:
    name_id: int
    platform_id: int
    plat_enc_id: int
    lang_id: int
    value: str


@dataclass
class SubsetSpec:
    mode: str = "preset"  # text | unicodes | preset
    text: str = ""
    unicodes: list[int] = field(default_factory=list)
    preset: str = "common_kr"
    layout_features: str = "*"


@dataclass
class BuildSpec:
    meta: Meta
    metrics: Metrics
    weights: list[WeightSpec]
    outputs: list[OutputSpec]
    extra_names: list[NameRecord]
    subset: SubsetSpec

    def to_dict(self) -> dict:
        return asdict(self)

    def enabled_outputs(self) -> list[str]:
        return [o.fmt for o in self.outputs if o.enabled]

    def find_output(self, fmt: str) -> Optional[OutputSpec]:
        for o in self.outputs:
            if o.fmt == fmt:
                return o
        return None

    def subset_unicodes(self) -> list[int]:
        s = self.subset
        if s.mode == "text":
            return sorted({ord(c) for c in s.text})
        if s.mode == "unicodes":
            return s.unicodes
        if s.mode == "preset":
            return _expand_preset(s.preset)
        raise ValueError(f"invalid subset.mode: {s.mode!r}")


# ── 파서 ────────────────────────────────────────────────────────

def _read_kv(ws) -> dict[str, str]:
    """key/value 시트 → dict. 1행 헤더 가정."""
    out: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        key = str(row[0]).strip()
        val = "" if row[1] is None else row[1]
        out[key] = val
    return out


def _read_metrics(ws) -> Metrics:
    m = Metrics()
    field_map = {
        ("head", "unitsperem"): "units_per_em",
        ("hhea", "ascent"): "hhea_ascent",
        ("hhea", "descent"): "hhea_descent",
        ("hhea", "linegap"): "hhea_line_gap",
        ("os/2", "stypoascender"): "typo_ascender",
        ("os/2", "stypodescender"): "typo_descender",
        ("os/2", "stypolinegap"): "typo_line_gap",
        ("os/2", "uswinascent"): "win_ascent",
        ("os/2", "uswindescent"): "win_descent",
        ("os/2", "ystrikeoutposition"): "strikeout_position",
        ("os/2", "ystrikeoutsize"): "strikeout_size",
        ("os/2", "fstype"): "fs_type",
        ("post", "underlineposition"): "underline_position",
        ("post", "underlinethickness"): "underline_thickness",
        ("os/2", "use_typo_metrics"): "use_typo_metrics",
    }
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None or row[1] is None:
            continue
        table = str(row[0]).strip().lower()
        fld = str(row[1]).strip().lower()
        attr = field_map.get((table, fld))
        if not attr:
            continue
        val = row[2]
        if val is None or val == "":
            continue
        if attr == "use_typo_metrics":
            setattr(m, attr, _to_bool(val))
        else:
            setattr(m, attr, _to_int(val, field=f"{table}.{fld}"))
    return m


def _read_weights(ws) -> list[WeightSpec]:
    out: list[WeightSpec] = []
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    required = [
        "weight_class", "style_name", "is_bold", "is_italic",
        "korean_family", "latin_family", "psname",
        "fullname_suffix", "base_font_key",
    ]
    for col in required:
        if col not in idx:
            raise ValueError(f"weights 시트에 '{col}' 열이 없음")

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[idx["weight_class"]] in (None, ""):
            continue
        out.append(WeightSpec(
            weight_class=_to_int(row[idx["weight_class"]], field="weight_class"),
            style_name=str(row[idx["style_name"]] or "").strip(),
            is_bold=_to_bool(row[idx["is_bold"]]),
            is_italic=_to_bool(row[idx["is_italic"]]),
            korean_family=str(row[idx["korean_family"]] or "").strip(),
            latin_family=str(row[idx["latin_family"]] or "").strip(),
            psname=str(row[idx["psname"]] or "").strip(),
            fullname_suffix=str(row[idx["fullname_suffix"]] or "").strip(),
            base_font_key=str(row[idx["base_font_key"]] or "").strip(),
            notes=str(row[idx["notes"]] or "").strip() if "notes" in idx else "",
        ))
    if not out:
        raise ValueError("weights 시트에 유효한 행이 없음")
    return out


def _read_outputs(ws) -> list[OutputSpec]:
    out: list[OutputSpec] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        fmt = str(row[0]).strip()
        if not fmt:
            continue
        out.append(OutputSpec(
            fmt=fmt,
            enabled=_to_bool(row[1]),
            psname_suffix=str(row[2] or "").strip() if len(row) > 2 else "",
        ))
    return out


def _read_names(ws) -> list[NameRecord]:
    out: list[NameRecord] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        # 마지막 안내 행("* nameID …") 같은 케이스 스킵
        try:
            nid = _to_int(row[0], field="nameID")
        except ValueError:
            continue
        platform = str(row[1] or "win").strip().lower()
        lang = str(row[2] or "en").strip().lower()
        value = "" if row[3] is None else str(row[3])
        if not value:
            continue
        if platform not in _PLATFORM_MAP:
            raise ValueError(f"names: 알 수 없는 platform {platform!r}")
        plat_id, enc_id = _PLATFORM_MAP[platform]
        if lang not in _LANG_MAP:
            raise ValueError(f"names: 알 수 없는 lang {lang!r}")
        out.append(NameRecord(
            name_id=nid,
            platform_id=plat_id,
            plat_enc_id=enc_id,
            lang_id=_LANG_MAP[lang],
            value=value,
        ))
    return out


def _read_subset(ws) -> SubsetSpec:
    kv = _read_kv(ws)
    mode = str(kv.get("mode", "preset")).strip().lower() or "preset"
    if mode not in ("text", "unicodes", "preset"):
        raise ValueError(f"subset.mode 는 text|unicodes|preset 중 하나 ({mode!r})")
    unicodes: list[int] = []
    if mode == "unicodes":
        spec = str(kv.get("unicodes", "")).strip()
        if not spec:
            raise ValueError("subset.mode=unicodes 인데 unicodes 셀이 비어있음")
        unicodes = _parse_unicode_ranges(spec)
    return SubsetSpec(
        mode=mode,
        text=str(kv.get("text", "") or ""),
        unicodes=unicodes,
        preset=str(kv.get("preset", "common_kr") or "common_kr").strip(),
        layout_features=str(kv.get("layout_features", "*") or "*").strip(),
    )


def parse(sheet_path: str | Path) -> BuildSpec:
    """XLSX 작업서를 BuildSpec 으로 파싱."""
    p = Path(sheet_path)
    if not p.exists():
        raise FileNotFoundError(p)
    wb = load_workbook(p, data_only=True)
    required = {"meta", "metrics", "weights", "outputs", "names", "subset"}
    missing = required - set(wb.sheetnames)
    if missing:
        raise ValueError(f"누락된 시트: {sorted(missing)}")

    meta_kv = _read_kv(wb["meta"])
    meta = Meta(
        project_name=str(meta_kv.get("project_name", "") or ""),
        version=str(meta_kv.get("version", "1.000") or "1.000"),
        author=str(meta_kv.get("author", "") or ""),
        client=str(meta_kv.get("client", "") or ""),
        vendor=str(meta_kv.get("vendor", "") or ""),
        designer_ko=str(meta_kv.get("designer_ko", "") or ""),
        designer_en=str(meta_kv.get("designer_en", "") or ""),
        vendor_url=str(meta_kv.get("vendor_url", "") or ""),
        designer_url=str(meta_kv.get("designer_url", "") or ""),
        license=str(meta_kv.get("license", "") or ""),
        license_url=str(meta_kv.get("license_url", "") or ""),
        copyright=str(meta_kv.get("copyright", "") or ""),
        trademark=str(meta_kv.get("trademark", "") or ""),
        name_record_scope=str(
            meta_kv.get("name_record_scope", "korean") or "korean"
        ).strip().lower(),
    )
    if meta.name_record_scope not in ("korean", "latin"):
        raise ValueError(
            f"meta.name_record_scope 는 korean|latin ({meta.name_record_scope!r})"
        )

    return BuildSpec(
        meta=meta,
        metrics=_read_metrics(wb["metrics"]),
        weights=_read_weights(wb["weights"]),
        outputs=_read_outputs(wb["outputs"]),
        extra_names=_read_names(wb["names"]),
        subset=_read_subset(wb["subset"]),
    )
